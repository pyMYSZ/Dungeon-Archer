"""
Dungeon Archer
Author:    MYSZ - https://github.com/pyMYSZ
type:      game - pygame
version:   1.7.2
date:      29-02-2024

 >> NOTE
Game based on udemy course 'Complete Pygame Tutorial - Create a Dungeon Crawler'
https://www.udemy.com/course/pygame-dungeon-crawler/?kw=create+dungeon&src=sac
"""

# region 01 - Import data
import pygame
import config.py.game_settings as gs
import config.py.functions as hf
from config.py.items import Item
from config.py.mobs import EnemyMob
from config.py.weapon import Floor
from config.py.weapon import Bow
from config.py.sprite_image import SpriteImage
from pygame.locals import *
from config.py import mobs
import csv
import math
import ctypes
import datetime
import openpyxl
import os
import webbrowser
import random

# endregion


# region 02 - Make a Display
# Screen settings
pygame.init()
display_surface = pygame.display.set_mode((gs.WINDOW_WIDTH, gs.WINDOW_HEIGHT))
pygame.display.set_caption(gs.GAME_NAME)


# Moving the game screen to the center of the monitor in Windows
def center_window():
    user32 = ctypes.windll.user32
    screen_center_x = user32.GetSystemMetrics(0) // 2
    screen_center_y = user32.GetSystemMetrics(1) // 2
    user32.SetWindowPos(pygame.display.get_wm_info()["window"], 0, screen_center_x - gs.WINDOW_WIDTH // 2,
                        screen_center_y - gs.WINDOW_HEIGHT // 2 - 50, 0, 0, 0x0001)


# discord channel invite system
def open_discord_channel():
    global discord_link_activ
    webbrowser.open("discord.gg/um6zuuKhxG") if discord_link_activ else None


# Game cursor
cursor_img = hf.scale_image(pygame.image.load("./config/images/game/cursor_ch.png"), 0.75)
pygame.mouse.set_visible(False)


# draw new clear display
def clear_display(show_info=True):
    display_surface.fill('black')
    if show_info:
        display_surface.blit(mysz_games_img, mysz_games_rect)
        display_surface.blit(discord_img, discord_rect)


center_window()

# Game Logo
logo_image = pygame.image.load('./config/images/game/archer.png')
pygame.display.set_icon(logo_image)

# endregion


# region 03 - Game Settings
# Background music
pygame.mixer.init()
pygame.mixer.music.load('./config/music/Dungeon-Crawler.wav')
pygame.mixer_music.play(-1, 0.0, 7500)
pygame.mixer.music.set_volume(0.2)

# Load Score Table
excel_file_path = './config/settings/score.xlsx'
if os.path.exists(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Score", "Date", "Weapon"])
    wb.save(excel_file_path)

# Game variable
discord_link_activ = True
FPS = gs.FPS_GAME
save_score = True
next_map = False
shop_map = False
start_intro = True
intro_running = True
load_bar_width = 0
menu_level_select = False
menu_weapon_select = False
main_menu = False
menu_about = False
menu_how_to_play = False
menu_HTP_index = 0
menu_mob_info = False
menu_mob_index = 0
mob_img_index = 0
mob_img_pos_x = 0
menu_item_info = False
menu_item_index = 0
menu_weapon_info = False
menu_weapon_info_index = 0
back_to_menu = False
GameOverText = False
menu_intro = False
menu_score = False
menu_level_info = ''
map_level = gs.get_value('max_level')
weapon_index = gs.get_value('weapon')
high_score = gs.get_value('high_score')
game_running = True
game_start = False
draw_menu_variable = True
weapon_menu_sound = True
weapon_info = False
player_info = False
back_click = False
restart_score = False
show_top11 = True
current_menu = "level"
selected_level = 1
menu_option = map_level - 1
coin_index = 0
timer = 0
unlocked_level = gs.get_value('max_level')
max_level = 40
music_level = 0  # 0 - 50%, 1 - 100%, 2 - OFF

screen_scroll = [0, 0]  # [x, y] scroll variable

# Sounds settings
sound_next = pygame.mixer.Sound('./config/music/click.mp3')
sound_reload = pygame.mixer.Sound('./config/music/reload.mp3')
sound_shop = pygame.mixer.Sound('./config/music/shop.mp3')

# Set FPS and clock
clock = pygame.time.Clock()

# Font settings
font_title = pygame.font.SysFont('Chiller', 75)
font_big = pygame.font.SysFont('Chiller', 200)
font_menu_big = pygame.font.SysFont('Chiller', 140)
font_menu_small = pygame.font.SysFont('Chiller', 50)
font_menu_item = pygame.font.SysFont('Chiller', 45)
font_info = pygame.font.SysFont('Chiller', 30)
font_mob_hp = pygame.font.SysFont('Chiller', 20)

# Images settings
intro_img = pygame.image.load('./config/images/intro/bg03.png')
intro_rect = intro_img.get_rect(bottomright=(gs.WINDOW_WIDTH, gs.WINDOW_HEIGHT))

mysz_games_img = pygame.image.load('./config/images/menu/mysz_games.png')
mysz_games_rect = mysz_games_img.get_rect(bottomright=(gs.WINDOW_WIDTH - 5, gs.WINDOW_HEIGHT - 5))

discord_img = hf.scale_image(pygame.image.load('./config/images/menu/discord.png'), 0.55)
discord_rect = discord_img.get_rect(bottomleft=(10, gs.WINDOW_HEIGHT - 10))

keyboard_img = pygame.image.load('./config/images/menu/settings.png')
archer_img = hf.scale_image(pygame.image.load('./config/images/menu/archer.png'), 0.70)
high_score_img = hf.scale_image(pygame.image.load('./config/images/menu/score.png'), 1.00)
menu_about_img = hf.scale_image(pygame.image.load('./config/images/menu/menu.png'), 1.00)

weapon_bow1_image = hf.scale_size_image(pygame.image.load('./config/images/hud/red_arrow.png'), 150, 150)
weapon_bow2_image = hf.scale_size_image(pygame.image.load('./config/images/hud/green_arrow.jpg'), 150, 150)
weapon_bow3_image = hf.scale_size_image(pygame.image.load('./config/images/hud/blue_arrow.jpg'), 150, 150)
weapon_bow4_image = hf.scale_size_image(pygame.image.load('./config/images/hud/orange_arrow.jpg'), 150, 150)
images_bow_list = [weapon_bow1_image, weapon_bow2_image, weapon_bow3_image, weapon_bow4_image]

# Load button images
start_img = pygame.image.load('./config/images/menu/button_start.png').convert_alpha()
exit_img = pygame.image.load('./config/images/menu/button_exit.png').convert_alpha()
back_img = pygame.image.load('./config/images/menu/button_back.png').convert_alpha()
about_img = pygame.image.load('./config/images/menu/button_about.png').convert_alpha()
score_img = pygame.image.load('./config/images/menu/button_score.png').convert_alpha()
about_game_img = pygame.image.load('./config/images/menu/game_info.png').convert_alpha()
about_enemy_img = pygame.image.load('./config/images/menu/mob_info.png').convert_alpha()
about_item_img = pygame.image.load('./config/images/menu/item_info.png').convert_alpha()
about_weapon_img = pygame.image.load('./config/images/menu/weapon_info.png').convert_alpha()
restart_score_img = pygame.image.load('./config/images/menu/restart_score_topLst.png').convert_alpha()
next_img = pygame.image.load('./config/images/menu/next.png').convert_alpha()
prev_img = pygame.image.load('./config/images/menu/prev.png').convert_alpha()

# Game Over
GAMEOVER_text = font_big.render('GAME OVER', True, gs.RED)
GAMEOVER_rect = GAMEOVER_text.get_rect(midleft=(150, gs.WINDOW_HEIGHT // 2 - 200))

game_over_text = font_title.render("Final Score: 2500", True, gs.WHITE)
game_over_rect = game_over_text.get_rect(center=(gs.WINDOW_WIDTH // 2 - 40, gs.WINDOW_HEIGHT // 2 - 15))
# endregion


# region 04 - Load and draw Map
# load tile map images
tile_list = []
for x in range(gs.TILE_AMOUNT + 1):
    tile_image = hf.scale_size_image(pygame.image.load(f'./config/images/maps/{x}.png').convert_alpha(),
                                     gs.TILE_SIZE, gs.TILE_SIZE)
    tile_list.append(tile_image)

# create empty tile list
world_data = []
for row in range(gs.ROW):
    r = [-1] * gs.COLUMNS
    world_data.append(r)

# load map data from csv file - load shop
with open(f'./config/maps/shop_map.csv', newline="") as csvfile:
    reader = csv.reader(csvfile, delimiter=',')
    for x, row in enumerate(reader):
        for y, tile in enumerate(row):
            if 0 <= x < len(world_data) and 0 <= y < len(world_data[x]):
                world_data[x][y] = int(tile)
            else:
                print(f'Error! x:{x}, y:{y} not defined')


# endregion


# region 05 - Screen Fade and Music Player
def change_music_level():
    global music_level
    music_level = (music_level + 1) % 3
    if music_level == 0:
        pygame.mixer.music.set_volume(0.12)
    elif music_level == 1:
        pygame.mixer.music.set_volume(0.22)
    elif music_level == 2:
        pygame.mixer.music.set_volume(0)


class ScreenFade:
    def __init__(self, colour, speed):
        self.colour = colour
        self.speed = speed
        self.counter = 0

    def fade(self):
        self.counter += self.speed
        pygame.draw.rect(display_surface, self.colour, (0 - self.counter, 0, gs.WINDOW_WIDTH // 2, gs.WINDOW_HEIGHT))
        pygame.draw.rect(display_surface, self.colour, (gs.WINDOW_WIDTH // 2 + self.counter, 0,
                                                        gs.WINDOW_WIDTH // 2, gs.WINDOW_HEIGHT))
        pygame.draw.rect(display_surface, self.colour, (0, 0 - self.counter, gs.WINDOW_WIDTH, gs.WINDOW_HEIGHT // 2))
        pygame.draw.rect(display_surface, self.colour, (0, gs.WINDOW_HEIGHT // 2 + self.counter, gs.WINDOW_WIDTH,
                                                        gs.WINDOW_HEIGHT // 2))

        stop_fade = False
        if self.counter >= max(gs.WINDOW_WIDTH // 2, gs.WINDOW_HEIGHT // 2):
            stop_fade = True
        return stop_fade


# endregion


# region 06 - Data
mob_types = ['elf', 'imp', 'skeleton', 'goblin', 'muddy', 'tiny_zombie', 'big_demon', 'big_ogre', 'big_zombie', 'slug']

animation_types = ['idle', 'run']

weapon_title_list = ['Damage:', 'Distance:', 'Cooldown:', 'Reload cd:', 'Amount:',
                     'Critical:', 'Slow:', 'Armor Pen.:']

weapon_title_list_full = ['Damage:', 'Distance:', 'Cooldown:', 'Reload cooldown:', 'Amount:', 'Critical Chance:',
                          'Slow Chance:', 'Armor Penetration:', 'Arrow acceleration:', 'Free Auto-shooter:']

weapon_value_list = [[' 60-85', '250-999', '0.38 sek.', '1.25 sek.', ' 3 x 10', '  30 %', '   0 %',
                      '  35 %', '  -2.50', '    NO'],
                     [' 25-65', ' 50-750', '0.28 sek.', '0.67 sek.', ' 4 x 25', '  10 %', '   5 %',
                      '  15 %', '  -0.50', '    NO'],
                     [' 20-35', '350-600', '0.20 sek.', '1.00 sek.', ' 5 x 30', '   5 %', '  25 %',
                      '   0 %', '  +1.80', '    NO'],
                     [' 9-14', '125-480', '0.09 sek.', '2.00 sek.', ' 4 x 55', '   1 %', '   1 %',
                      '  20 %', '   1.00', ' YES (max 1.25 sek.)']]

weapon_info_list_full = [['The damage value is reduced by the enemy armor',
                          'Damage = 100 / (100 + ARMOR * (100-armor.penetration) / 100)'],
                         ['You achieve maximum DMG at minimum distance and vice versa'],
                         [f'1 Attack every {weapon_value_list[weapon_index][2]}', 'Attack Speed = 1000 / cooldown'],
                         ['Time of loading new ammunition'], ['Initial amount of arrows'],
                         ['Raises damage to DMG + DMG*(100 - crit.reduction) / 100'],
                         ['Chance to slow down opponent on hit '],
                         ['Reduces the armor level of the enemy',
                          'Damage = 100 / (100 + ARMOR * (100-armor.penetration) / 100)'],
                         ['Arrows accelerate(+)/decelerate(-) during the flight by this factor'],
                         ['Automatically shoot multiple shots at maximum speed with one click']]

HTP_info_text_list = [['Player movement: W, A, S, D keys',
                       'Hold SPACE to increase speed (if you have energy)',
                       'Shooting left mouse button'],
                      ['Press "R" to buy reload weapon',
                       'Press "B" to buy extra ammunition (5 coins)',
                       'Press "Q" to activate auto-shooting mode (1 sec = 1 coin)'],
                      ['Press "i" to enable/disable weapon stats info',
                       'Press "o" to enable/disable player stats info',
                       'Press "C" to enable/disable both info'],
                      ['Press "L" to enable/disable level info',
                       "If the game doesn't refresh, press 'P'"],
                      ['Use the brown ladder to go to the next level',
                       'Use the green ladder to use the STORE before the next level'],
                      ['In the store, get close to the $ sign and press "E" to move on',
                       'Cost 4-6 coins. You must have at least 5 coins to use']]

mob_info_text_list = ['Health:', 'Armor:', 'Speed:', 'Damage:', 'Attack Range:', 'Attack Cooldown:', 'Crit. Reduction:']

mob_extra_info_list = [['Attacks with fireballs from a distance',
                        'Fireball damage is reduced [%] by your fire resist value.',
                        '5% chance of dropping a coin, arrow box, red or blue potion'],
                       ['Monster dies after dealing you damage',
                        "Monster's attacks reduce your armor"],
                       ["Monster's attacks reduce your speed",
                        'Player regenerates his speed after drinking blue potion',
                        '10% chance of dropping a coin, arrow box, red or blue potion'],
                       ["Monster's attacks steal player's energies",
                        "25% chance of dropping a coin, arrow box, red or blue potion"],
                       ['Monster dies after dealing you damage',
                        "Monster's attacks reduce your armor",
                        '5% chance of dropping a coin, arrow box, red or blue potion'],
                       ['Attacks with fireballs from a distance',
                        'Fireball damage is reduced [%] by your fire resist value.',
                        'Drop random item (coin, arrow box, potion) after die'],
                       ["Monster's attacks reduce your speed",
                        'Player regenerates his speed after drinking blue potion',
                        'Drop random potion (hp, mana, green, orange) after die',
                        'Chance of green / orange potion is twice as high'],
                       ['BOSS regenerates his health',
                        'Drop random potion (from shop) after die',
                        'Each monster attack reduces the amount of Coins by 1'],
                       ["Monster's attacks reduce your speed",
                        'Player regenerates his speed after drinking blue potion',
                        '10% chance of dropping a coin, arrow box, red or blue potion']]

mob_stats_text_list = []
for i in range(len(mobs.armor_list)):
    mob_stats_list_index_i = [str(mobs.health_list[i]), str(mobs.armor_list[i]),
                              str(mobs.velocity_list[i]),
                              str(f'{mobs.min_damage_list[i]} - {mobs.max_damage_list[i]}'),
                              str(mobs.attack_range_list[i]), str(mobs.attack_cooldown_list[i]),
                              str(f'{mobs.crit_reduction_list[i]} %')]
    mob_stats_text_list.append(mob_stats_list_index_i)

potion_info_text_list = [['Player max HP + 10-15', ' Player HP +15'], ['Player velocity + 0.15-0.20'],
                         ['Slow chance + 3%', 'Arrow acceleration + 0.45', 'Min DMG + 1'],
                         ['Attack speed +10%'], ['Fireball resist + 10-13%'],
                         ['Armor + 15-20', 'Player max HP + 3-5,  Player HP +5'],
                         ['Min DMG + 2-4', 'Max DMG + 3+5'], ['Critical strike chance + 3%', 'Min DMG + max(+1 or 4%)'],
                         ['Armor penetration + 3-5%', 'Max DMG + max(+1 or 5%)'], ['Player health + 12-15'],
                         ['Player energy  + 20%', 'Removes slowdown'],
                         ["Increases player's speed by 0.10", 'Increases arrow acceleration by 0.85'],
                         ['Increases maximum number of arrows in eq by 10%', 'Increases arrow acceleration by 0.25']]


# endregion


# 07 - Classes
class SpriteImage:
    # Sprite image in pygame
    def __init__(self, image):
        self.sheet = image
        self.width = self.sheet.get_width()
        self.height = self.sheet.get_height()

    def get_image(self, frame, width, height, scale, colour):
        g_image = pygame.Surface((width, height)).convert_alpha()
        g_image.blit(self.sheet, (0, 0), ((frame * width), 0, width, height))
        g_image = pygame.transform.scale(g_image, (width * scale, height * scale))
        g_image.set_colorkey(colour)  # make image transparent

        return g_image

    def get_image_one_sheet(self, amount, index=1, scale=1, x=0, y=0):
        width = self.width // amount
        height = self.height
        image = pygame.Surface((width, height)).convert_alpha()
        image.blit(self.sheet, (x, y), ((index * width), 0, width, height))
        image = pygame.transform.scale(image, (width * scale, height * scale))
        image.set_colorkey('black')  # make image transparent
        # color_key = image.get_at((0, 0))
        # image = pygame.transform.scale(image, (width * scale, height * scale))
        # image.set_colorkey(color_key)
        return image


class Button:
    def __init__(self, x, y, image, index=0, animation=False):
        self.image_list = image if animation else [image]
        self.index = index
        self.image = self.image_list[self.index]
        self.rect = self.image.get_rect(topleft=(x, y))

    def draw(self, surface):
        action = False
        pos = pygame.mouse.get_pos()

        if self.rect.collidepoint(pos):
            if pygame.mouse.get_pressed()[0]:
                action = True
        self.image.set_colorkey('black')
        surface.blit(self.image, self.rect)

        return action

    def update(self):
        self.index = (self.index + 1) % 3
        self.image = self.image_list[self.index]


class Map:
    def __init__(self):
        self.map_elements = []
        self.obstacle_list = []
        self.teleport_list = []
        self.floor_list = []
        self.away_world_list = []  # DEBUG
        self.item_list = []
        self.mob_list = []
        self.shop_list = []

        self.player = None
        self.next_map = None
        self.next_level = None
        self.shop_map = None

    def data(self, tile_list, data, items_images, characters_list):
        for y, row in enumerate(data):
            for x, tile in enumerate(row):
                # print(f"Debug: {y}, {x}, {tile}")
                image = tile_list[tile]
                image_x = x * gs.TILE_SIZE
                image_y = y * gs.TILE_SIZE
                image_rect = image.get_rect(center=(image_x, image_y))
                tile_data = [image, image_rect, image_x, image_y]

                # add image data to main title list
                if tile == 5:     # next map
                    self.next_map = tile_data
                if tile == 6:     # player start
                    player = Player(image_x, image_y, characters_list[0])
                    self.player = player
                    # tile_data[0] = tile_list[31]
                if tile == 7:  # wall
                    self.obstacle_list.append(tile_data)
                if tile == 10:     # HP potion
                    hp_potion = Item(x=image_x, y=image_y, typ=1, image_list=items_images)
                    self.item_list.append(hp_potion)
                    tile_data[0] = tile_list[0]
                if tile == 11:     # mana potion
                    mana_potion = Item(x=image_x, y=image_y, typ=2, image_list=items_images)
                    self.item_list.append(mana_potion)
                    tile_data[0] = tile_list[0]
                if tile == 12:     # arrow
                    arrow_box = Item(x=image_x, y=image_y, typ=3, image_list=items_images)
                    self.item_list.append(arrow_box)
                    tile_data[0] = tile_list[0]
                if tile == 13:     # coin
                    coin = Item(x=image_x, y=image_y, typ=0, image_list=items_images)
                    self.item_list.append(coin)
                    tile_data[0] = tile_list[0]
                if tile == 14:     # mob typ(index) 5 = Big demon boss
                    mob_05 = EnemyMob(image_x, image_y, 5, characters_list, True)
                    self.mob_list.append(mob_05)
                    tile_data[0] = tile_list[0]
                if tile == 15:     # mob typ(index) 0
                    mob_00 = EnemyMob(image_x, image_y, 0, characters_list)
                    self.mob_list.append(mob_00)
                    tile_data[0] = tile_list[0]
                if tile == 16:     # mob typ(index) 1
                    mob_01 = EnemyMob(image_x, image_y, 1, characters_list, typ='armor')
                    self.mob_list.append(mob_01)
                    tile_data[0] = tile_list[0]
                if tile == 17:     # mob typ(index) 2
                    mob_02 = EnemyMob(image_x, image_y, 2, characters_list, typ='slow')
                    self.mob_list.append(mob_02)
                    tile_data[0] = tile_list[0]
                if tile == 18:     # mob typ(index) 3
                    mob_03 = EnemyMob(image_x, image_y, 3, characters_list, typ='boost')
                    self.mob_list.append(mob_03)
                    tile_data[0] = tile_list[0]
                if tile == 19:     # mob typ(index) 4
                    mob_04 = EnemyMob(image_x, image_y, 4, characters_list, typ='armor')
                    self.mob_list.append(mob_04)
                    tile_data[0] = tile_list[0]
                if tile == 20:  # wall info green
                    self.obstacle_list.append(tile_data)
                if tile == 21:  # wall info yellow
                    self.obstacle_list.append(tile_data)
                if tile == 22:  # wall info red
                    self.obstacle_list.append(tile_data)
                if tile == 23:  # floor slow
                    floor = Floor(image_x, image_y, animation_cooldown=700)
                    self.floor_list.append(floor)
                if tile == 24:  # flor fast
                    floor = Floor(image_x, image_y, animation_cooldown=500)
                    self.floor_list.append(floor)
                if tile == 25:  # tp wall right
                    self.teleport_list.append(tile_data)
                if tile == 26:  # tp wall down
                    self.teleport_list.append(tile_data)
                if tile == 27:  # mob typ(index) 6 = Big Ogr boss
                    mob_06 = EnemyMob(image_x, image_y, 6, characters_list, True, typ='slow')
                    self.mob_list.append(mob_06)
                    tile_data[0] = tile_list[0]
                if tile == 28:  # mob typ(index) 7 = Big Zombie boss
                    mob_07 = EnemyMob(image_x, image_y, 7, characters_list, True, typ='Zombie Boss')
                    self.mob_list.append(mob_07)
                    tile_data[0] = tile_list[0]
                if tile == 29:  # mob typ(index) 8 = Slug
                    mob_08 = EnemyMob(image_x, image_y, 8, characters_list, False, typ='slow')
                    self.mob_list.append(mob_08)
                    tile_data[0] = tile_list[0]
                if tile == 30:  # shop coin system
                    self.shop_list.append(tile_data)
                if tile == 31:  # shop map
                    self.shop_map = tile_data
                if tile == 32:  # SHOP potion 1
                    hp_potion = Item(x=image_x, y=image_y, typ=11, image_list=items_images)
                    self.item_list.append(hp_potion)
                    tile_data[0] = tile_list[0]
                if tile == 33:  # SHOP potion 2
                    hp_potion = Item(x=image_x, y=image_y, typ=12, image_list=items_images)
                    self.item_list.append(hp_potion)
                    tile_data[0] = tile_list[0]
                if tile == 34:  # SHOP potion 3
                    hp_potion = Item(x=image_x, y=image_y, typ=13, image_list=items_images)
                    self.item_list.append(hp_potion)
                    tile_data[0] = tile_list[0]
                if tile == 35:  # SHOP potion 4
                    hp_potion = Item(x=image_x, y=image_y, typ=14, image_list=items_images)
                    self.item_list.append(hp_potion)
                    tile_data[0] = tile_list[0]
                if tile == 36:  # SHOP potion 5
                    hp_potion = Item(x=image_x, y=image_y, typ=15, image_list=items_images)
                    self.item_list.append(hp_potion)
                    tile_data[0] = tile_list[0]
                if tile == 37:  # SHOP potion 6
                    hp_potion = Item(x=image_x, y=image_y, typ=16, image_list=items_images)
                    self.item_list.append(hp_potion)
                    tile_data[0] = tile_list[0]
                if tile == 38:  # SHOP potion 7
                    hp_potion = Item(x=image_x, y=image_y, typ=17, image_list=items_images)
                    self.item_list.append(hp_potion)
                    tile_data[0] = tile_list[0]
                if tile == 39:  # SHOP potion 8
                    hp_potion = Item(x=image_x, y=image_y, typ=18, image_list=items_images)
                    self.item_list.append(hp_potion)
                    tile_data[0] = tile_list[0]
                if tile == 40:  # SHOP potion 9
                    hp_potion = Item(x=image_x, y=image_y, typ=19, image_list=items_images)
                    self.item_list.append(hp_potion)
                    tile_data[0] = tile_list[0]

                if tile >= 0:    # cuz empty places is -1
                    self.map_elements.append(tile_data)

    def draw(self, surface):
        for tile in self.map_elements:
            surface.blit(tile[0], tile[1])

    def update(self, screen_scroll):
        for tile in self.map_elements:
            # tile = [image, image_rect, image_x, image_y]
            tile[2] += screen_scroll[0]
            tile[3] += screen_scroll[1]
            tile[1].center = (tile[2], tile[3])


class Player:
    def __init__(self, x, y, image_list):
        self.update_time = pygame.time.get_ticks()
        self.velocity = 0
        self.velocity_normal = 0
        self.velocity_scale = gs.PLAYER_BOOST_SCALE
        self.max_hp = 100     # maximum health
        self.health = 100     # start health
        self.armor = 100      # reduces damage from opponents by (100 / 100+armor) times
        self.fire_resist = 0  # fire resistance (0-100%)
        self.boost_lvl = 750  # start boost level
        self.dead = False
        self.slowed = False
        self.show_level = False
        self.map_level = 0

        self.score = 0
        self.coin_amount = 0
        self.max_boost = 2000
        self.animation_cooldown = 70
        self.weapon_index = 0
        self.last_reload = pygame.time.get_ticks()  # reload cooldown

        # images setting
        self.rect = pygame.Rect(0, 0, 32, 40)
        self.rect.center = (x, y)

        self.flip_x = False
        self.flip_y = False
        self.image_index = 0
        self.image_type = 0  # 0 = player idle, 1 = player running
        self.image_list = image_list
        self.image = self.image_list[self.image_type][self.image_index]
        self.hud_img_index = 0

        # Move settings
        self.horizontal = False
        self.vertical = False
        self.shoping = False
        self.move_up = True
        self.image_type = 0
        self.dx = 0
        self.dy = 0

        # mob AI system
        self.rect2 = pygame.Rect(0, 0, 32, 40)
        self.rect3 = pygame.Rect(0, 0, 32, 40)
        self.counter = 0
        self.temp_counter = 0

    def update(self):
        self.counter += 1
        # make animation
        self.image = self.image_list[self.image_type][self.image_index]

        # change images - make animation
        if pygame.time.get_ticks() - self.update_time > self.animation_cooldown:
            self.image_index = (self.image_index + 1) % 4
            self.update_time = pygame.time.get_ticks()

        # reset mana bug
        if self.boost_lvl > self.max_boost:
            self.boost_lvl = self.max_boost

        # check player die
        if self.health <= 0:
            self.dead = True

        # mob AI system
        if self.counter >= gs.AI_COUNTER:
            # print(f'change rect {self.temp_counter}')
            self.counter = 0
            self.temp_counter += 1
            self.rect2.centerx = self.rect.centerx + random.randint(-15, 15)
            self.rect2.centery = self.rect.centery + random.randint(-10, 10)

            self.rect3.centerx = self.rect.centerx + random.randint(-100, 100)
            self.rect3.centery = self.rect.centery + random.randint(-75, 75)

    def draw(self, surface):
        flipped_image = pygame.transform.flip(self.image, self.flip_x, self.flip_y)
        surface.blit(flipped_image, (self.rect.x - 5, self.rect.y - gs.PLAYER_OFFSET * gs.SCALE))
        # pygame.draw.rect(surface, gs.RED, self.rect, 1)

    def draw_hud(self, surface, weapon):
        # draw hud
        pygame.draw.rect(surface, gs.GRAY, (0, 0, gs.WINDOW_WIDTH, gs.HUD))
        pygame.draw.line(surface, gs.WHITE, (0, gs.HUD), (gs.WINDOW_WIDTH, gs.HUD), 2)

        # draw coin info
        coin_text = f'{self.coin_amount:02}'
        hf.draw_text(surface, coin_text, 280, 5, gs.font_hud)

        # draw arrow info
        arrow_image = hf.scale_image(pygame.image.load(f'./config/images/hud/arrow.png').convert_alpha(), 0.9)
        hf.draw_image(surface, arrow_image, 350, 5)
        arrow_text = f'{weapon.amount:02}/{weapon.max_amount:02}'
        hf.draw_text(surface, arrow_text, 390, 5, gs.font_hud_num)

        box_image = hf.scale_image(pygame.image.load(f'./config/images/hud/bow_box.png').convert_alpha(), 1.25)
        for i in range(weapon.storage):
            hf.draw_image(surface, box_image, 500 + i*20, 5)

        # draw score info
        score_text = f'SCORE: {self.score:03}'
        hf.draw_text(surface, score_text, gs.WINDOW_WIDTH-250, 5, gs.font_hud)

        # draw level info:
        level_text = f'Level: {self.map_level}' if self.show_level else ''
        hf.draw_text(surface, level_text, gs.WINDOW_WIDTH-440, 5, gs.font_hud)

        # draw player (Kenny) images
        player_images_list = []
        player_img = pygame.image.load(f'./config/images/hud/kenny.png')
        player_image = SpriteImage(player_img)

        for i in range(0, 10):
            image = player_image.get_image_one_sheet(10, i, 1)
            player_images_list.append(image)

        del player_images_list[1]
        del player_images_list[1]

        if self.health > 0:
            self.hud_img_index = 7 - self.health // 15
        else:
            self.dead = True

        hf.draw_image(surface, player_images_list[self.hud_img_index], 0, 0)

        # draw hp and boost bars info
        bar_image = pygame.image.load(f'./config/images/hud/bar.png').convert_alpha()
        bar_width = bar_image.get_width() - 1
        ratio = self.boost_lvl / self.max_boost
        hp_ratio = self.health / self.max_hp
        pygame.draw.rect(surface, gs.BAR_END, (48, 35, bar_width, 15))
        pygame.draw.rect(surface, gs.BAR_TOP, (48, 36, bar_width * ratio, 11))
        pygame.draw.rect(surface, gs.RED, (48, 0, bar_width, 35))
        pygame.draw.rect(surface, gs.BAR_END, (48, 4, bar_width * hp_ratio, 30))
        hf.draw_image(surface, bar_image, 48, 0)

    def move(self, weapon=None, obstacle_list=None):
        """ Update the player position after move """
        if obstacle_list is None:
            obstacle_list = []
        self.horizontal = False
        self.vertical = False
        self.shoping = False
        self.move_up = True
        self.image_type = 0
        self.dx = 0
        self.dy = 0

        keys = pygame.key.get_pressed()
        if keys[pygame.K_a] and self.rect.left > gs.PLAYER_OFFSET:
            self.horizontal = True
            self.flip_x = True
            self.image_type = 1
            self.dx -= self.velocity / math.sqrt(self.horizontal + self.vertical)
        if keys[pygame.K_d] and self.rect.right < gs.WINDOW_WIDTH - gs.PLAYER_OFFSET:
            self.horizontal = True
            self.flip_x = False
            self.image_type = 1
            self.dx += self.velocity / math.sqrt(self.horizontal + self.vertical)
        if keys[pygame.K_w] and self.rect.top > gs.HUD:
            self.vertical = True
            self.move_up = True
            self.image_type = 1
            self.dy -= self.velocity / math.sqrt(self.horizontal + self.vertical)
        if keys[pygame.K_s] and self.rect.bottom < gs.WINDOW_HEIGHT - gs.PLAYER_OFFSET:
            self.vertical = True
            self.move_up = False
            self.image_type = 1
            self.dy += self.velocity / math.sqrt(self.horizontal + self.vertical)

        # check for collision in horizontal direction
        self.rect.x += self.dx
        for obstacle in obstacle_list:
            # obstacle / next_map / next_level = [image, rect, x, y]
            if obstacle[1].colliderect(self.rect):
                if self.dx < 0:
                    self.rect.left = obstacle[1].right
                if self.dx > 0:
                    self.rect.right = obstacle[1].left

        # check for collision in vertical direction
        self.rect.y += self.dy
        for obstacle in obstacle_list:
            if obstacle[1].colliderect(self.rect):
                if self.dy < 0:
                    self.rect.top = obstacle[1].bottom
                if self.dy > 0:
                    self.rect.bottom = obstacle[1].top

        # Engage Boost
        if keys[pygame.K_SPACE] and self.boost_lvl > 20:
            self.velocity = self.velocity_normal * self.velocity_scale
            self.boost_lvl -= 15
            self.animation_cooldown = 40
        else:
            self.velocity = self.velocity_normal
            self.animation_cooldown = 70

        # Used Shop
        if keys[pygame.K_e]:
            self.shoping = True

        # Show level
        if keys[pygame.K_l]:
            self.show_level = not self.show_level

        # Reload gun
        if keys[pygame.K_r] and weapon.storage > 0 and weapon.amount != weapon.max_amount:
            self.last_reload = pygame.time.get_ticks()
            weapon.amount = weapon.max_amount
            weapon.storage -= 1
            weapon.reloading = True
            sound_reload.play()

        if pygame.time.get_ticks() - self.last_reload > weapon.reload_cooldown:
            weapon.reloading = False

        # Boost regeneration
        if self.max_boost > self.boost_lvl + 1 and not self.slowed:
            if self.boost_lvl > 10:
                self.boost_lvl += 1
            else:
                self.boost_lvl += 0.1

        # Speed regeneration
        if not self.slowed and self.velocity_normal < gs.PLAYER_VELOCITY:
            # faster regeneration if player dont move
            if self.vertical or self.horizontal:
                self.velocity_normal += 0.00012
            else:
                self.velocity_normal += 0.00030

        # Debug speed
        if self.velocity_normal < 0.83:
            self.velocity_normal = 0.85
        if self.velocity_normal < 0.73:
            self.velocity_normal = 0.85

    def move_camera(self):
        screen_scroll = [0, 0]
        # Camera move left and right
        if self.rect.right > gs.WINDOW_WIDTH - gs.SCROLL_THRESH:
            screen_scroll[0] = gs.WINDOW_WIDTH - gs.SCROLL_THRESH - self.rect.right
            self.rect.right = gs.WINDOW_WIDTH - gs.SCROLL_THRESH
        elif self.rect.left < gs.SCROLL_THRESH:
            screen_scroll[0] = gs.SCROLL_THRESH - self.rect.left
            self.rect.left = gs.SCROLL_THRESH

        # Camera move up and down
        if self.rect.top < gs.SCROLL_THRESH:
            screen_scroll[1] = gs.SCROLL_THRESH - self.rect.top
            self.rect.top = gs.SCROLL_THRESH
        elif self.rect.bottom > gs.WINDOW_HEIGHT - gs.SCROLL_THRESH:
            screen_scroll[1] = gs.WINDOW_HEIGHT - gs.SCROLL_THRESH - self.rect.bottom
            self.rect.bottom = gs.WINDOW_HEIGHT - gs.SCROLL_THRESH
        return screen_scroll

    def obstacle_collision(self, next_map, teleport_list, shop_list, shop_map):
        # Teleport obstacle list
        for obstacle in teleport_list:
            if obstacle[1].colliderect(self.rect):
                if self.horizontal:
                    self.rect.left = obstacle[1].right
                if self.vertical:
                    self.rect.top = obstacle[1].bottom

        # Shop obstacle list
        for obstacle in shop_list:
            if obstacle[1].colliderect(self.rect):
                distance = hf.diagonal_distance(self.rect.centerx, obstacle[1].centerx,
                                                self.rect.centery, obstacle[1].centery)
                if self.shoping and self.coin_amount >= 5:
                    self.rect.left += 50
                    self.coin_amount -= random.randint(2, 3)
                    sound_shop.play()
                    self.coin_amount = 0 if self.coin_amount < 0 else self.coin_amount
                    self.shoping = False
                elif distance < 20:
                    self.rect.left -= 15
                elif distance < 15:
                    self.rect.left -= 10

        # check if player step in next map tile
        if next_map:
            if next_map[1].colliderect(self.rect):
                distance = hf.diagonal_distance(self.rect.centerx, next_map[1].centerx,
                                                self.rect.centery, next_map[1].centery)
                if distance < 10:
                    next_map = True

        # check if player step in shop map tile
        if shop_map:
            if shop_map[1].colliderect(self.rect):
                distance = hf.diagonal_distance(self.rect.centerx, shop_map[1].centerx,
                                                self.rect.centery, shop_map[1].centery)
                if distance < 15:
                    shop_map = True

        return next_map, shop_map

    def restart(self):
        self.velocity_normal = gs.PLAYER_VELOCITY
        self.score = 0
        self.coin_amount = 0
        self.dead = False
        self.max_hp = 100
        self.max_boost = 2000
        self.slowed = False

    def stats(self, bow_index):
        """ in progress """
        if bow_index == 0:
            self.velocity_normal = 5
            self.velocity_scale = 2
            self.armor = 125
            self.health = 100


# region 07 - Load and create objects
# Load coin images to HUD
coin_images = []
coin_img_sheet = pygame.image.load(f'./config/images/items/coin-blog.png')
coin_image = SpriteImage(coin_img_sheet)
for i in range(0, 6):
    image = coin_image.get_image_one_sheet(amount=6, index=i, scale=0.31)
    coin_images.append(image)

# Load character images
characters_list = []
mob_menu_list = []
for mob_type in mob_types:
    mob_image_list = []
    mob_menu_image_list = []
    for animation_typ in animation_types:
        temporary_list = []
        temporary_menu_list = []
        for i in range(4):
            # GAME mobs list
            image = pygame.image.load(f'./config/images/characters/{mob_type}/{animation_typ}/{i}.png').convert_alpha()
            image = hf.scale_image(image, gs.SCALE)
            temporary_list.append(image)
            # MENU mobs list
            image_size = 180 if mob_type.startswith('big') else 75
            mob_menu_image = hf.scale_size_image(image, image_size, image_size)
            temporary_menu_list.append(mob_menu_image)

        mob_image_list.append(temporary_list)
        mob_menu_image_list.append(temporary_menu_list)

    characters_list.append(mob_image_list)
    mob_menu_list.append(mob_menu_image_list)

# Load HowToPlay menu images
menu_how_to_play_images_list = []
for i in range(6):
    image = pygame.image.load(f'./config/images/menu/how_to_play/0{i}.png').convert_alpha()
    menu_how_to_play_images_list.append(image)

# Load potion images to 'About Menu / Item info'
potion_images_list = []
for i in range(9):
    image = pygame.image.load(f'./config/images/menu/item_info/menu_item_{i}.png').convert_alpha()
    image = hf.scale_image(image, 0.3)
    potion_images_list.append(image)

for i in range(4):
    image = pygame.image.load(f'./config/images/menu/item_info/potion_{i}.png').convert_alpha()
    potion_images_list.append(image)

# Load music images
music_images = []
music_img_sheet = pygame.image.load(f'./config/images/game/music.png')
music_image = SpriteImage(music_img_sheet)
for i in range(0, 3):
    image = music_image.get_image_one_sheet(amount=3, index=i, scale=0.50)
    music_images.append(image)
# print(len(music_images))
# Create images list to import map.py  - item list create in file mobs.py
item_list = mobs.item_list

# Create player's Weapon
players_bow = Bow(index=weapon_index)

# Create sprite groups
arrow_group = pygame.sprite.Group()
damage_text_group = pygame.sprite.Group()
item_group = pygame.sprite.Group()
enemy_group = pygame.sprite.Group()
fireball_group = pygame.sprite.Group()
floor_group = pygame.sprite.Group()

# Create map
world = Map()
world.data(tile_list, world_data, item_list, characters_list)


# Load next tail map function
def load_next_map():
    # clear group
    arrow_group.empty()
    damage_text_group.empty()
    item_group.empty()
    fireball_group.empty()
    floor_group.empty()
    enemy_group.empty()

    # create empty tile list
    map_data = []
    for x_row in range(gs.ROW):
        x_r = [-1] * gs.COLUMNS
        map_data.append(x_r)
    return map_data


# Create player
player = world.player

# Create enemy (list)
enemy_list = world.mob_list
for enemy in enemy_list:
    enemy_group.add(enemy)

# Add items from map(level) data
for item in world.item_list:
    item_group.add(item)

# Add floor from map(level) data
for floor in world.floor_list:
    floor_group.add(floor)

# Create screen fade
intro_fade = ScreenFade(gs.RED, 7)
menu_fade = ScreenFade(gs.BLACK, 10)

# Create button
# game over menu
button_restart = Button(gs.WINDOW_WIDTH // 2 - 450, gs.WINDOW_HEIGHT // 2 + 150, start_img)
button_about = Button(gs.WINDOW_WIDTH // 2 - 150, gs.WINDOW_HEIGHT // 2 + 150, about_img)
button_score = Button(gs.WINDOW_WIDTH // 2 + 150, gs.WINDOW_HEIGHT // 2 + 150, score_img)
button_main_menu = Button(gs.WINDOW_WIDTH // 2 - 275, gs.WINDOW_HEIGHT // 2 + 300, back_img)
button_exit = Button(gs.WINDOW_WIDTH // 2 + 25, gs.WINDOW_HEIGHT // 2 + 300, exit_img)

# about menu
button_info_game = Button(gs.WINDOW_WIDTH // 2 - 250, gs.WINDOW_HEIGHT // 2 - 375, about_game_img)
button_info_mob = Button(gs.WINDOW_WIDTH // 2 - 250, gs.WINDOW_HEIGHT // 2 - 240, about_enemy_img)
button_info_item = Button(gs.WINDOW_WIDTH // 2 - 250, gs.WINDOW_HEIGHT // 2 - 105, about_item_img)
button_info_weapon = Button(gs.WINDOW_WIDTH // 2 - 250, gs.WINDOW_HEIGHT // 2 + 30, about_weapon_img)
button_info_back = Button(gs.WINDOW_WIDTH // 2 - 275, gs.WINDOW_HEIGHT // 2 + 300, back_img)
button_info_exit = Button(gs.WINDOW_WIDTH // 2 + 25, gs.WINDOW_HEIGHT // 2 + 300, exit_img)

# sub about menu
button_next_option = Button(gs.WINDOW_WIDTH // 2 + 360, gs.WINDOW_HEIGHT // 2 - 100, next_img)
button_prev_option = Button(gs.WINDOW_WIDTH // 2 - 465, gs.WINDOW_HEIGHT // 2 - 100, prev_img)
button_next_weapon = Button(1030, 95, hf.scale_image(next_img, 0.38))
button_prev_weapon = Button(980, 95, hf.scale_image(prev_img, 0.38))
button_about_back = Button(gs.WINDOW_WIDTH // 2 - 275, gs.WINDOW_HEIGHT // 2 + 300, back_img)

# score menu
button_score_back = Button(gs.WINDOW_WIDTH // 2 - 275, gs.WINDOW_HEIGHT // 2 + 300, back_img)
button_score_restart = Button(850, 590, hf.scale_image(restart_score_img, 0.60))
# button_score_backUP = Button(gs.WINDOW_WIDTH // 2 - 250, gs.WINDOW_HEIGHT // 2 - 200, 1)

# music button
button_music = Button(x=gs.WINDOW_WIDTH - 55, y=3, image=music_images, index=music_level, animation=True)
# endregion


# 08 Main Game Loop
while game_running:
    # region 8.1 Event handler
    for event in pygame.event.get():
        # DisplayGame X button
        if event.type == pygame.QUIT:
            game_running = False
        # Mouse click
        elif event.type == pygame.MOUSEBUTTONDOWN:
            # Open DC channel
            if discord_rect.collidepoint(event.pos):
                open_discord_channel()
            # Select Level
            elif menu_level_select:
                for i in range(max_level):
                    bounding_box = pygame.Rect(120 * (2 + i % 8) - 100, 70 + 120 * (1 + i // 8), 90, 90)
                    if bounding_box.collidepoint(event.pos):
                        menu_option = i
                        # Handle the selection logic here, similar to the keyboard input
                        if menu_option < unlocked_level:
                            map_level = int(menu_option + 1)
                            menu_level_select = False
                            menu_weapon_select = True
                            menu_option = players_bow.index if back_to_menu else weapon_index
                            current_menu = "weapon"
                        else:
                            menu_level_info = 'This level is not yet unlocked'
            # Select Weapon
            elif menu_weapon_select:
                for i in range(4):
                    if pygame.Rect(225 + 235 * i, 160, 205, 700).collidepoint(event.pos):
                        menu_option = i
                        # Handle the selection logic here, similar to the keyboard input
                        menu_weapon_select = False
                        weapon_menu_sound = False
                        gs.update_value('weapon', int(menu_option))
                        game_start = True
                        player.dead = True
                        main_menu = True
                        menu_about = False
                        players_bow.index = int(menu_option)
        # Keyboard click
        elif event.type == KEYDOWN:
            # Show info (weapon and player values)
            if event.key == K_i:
                weapon_info = not weapon_info
            if event.key == K_o:
                player_info = not player_info
            if event.key == K_c:
                if player_info == weapon_info:
                    player_info, weapon_info = not player_info, not weapon_info
                else:
                    player_info = weapon_info
            # FPS debug
            if event.key == K_p:
                FPS = gs.FPS_GAME
            # Skip intro
            if event.key == K_SPACE and intro_running:
                intro_running = False
                game_start = True
                player.dead = True
                main_menu = True
                pygame.mouse.set_visible(True)
            # Select Level
            if current_menu == "level":
                if event.key == K_w or event.key == K_UP:
                    sound_next.play()
                    menu_option = (menu_option - 8) % 40
                    menu_level_info = ''
                elif event.key == K_s or event.key == K_DOWN:
                    menu_option = (menu_option + 8) % 40
                    menu_level_info = ''
                    sound_next.play()
                elif event.key == K_a or event.key == K_LEFT:
                    sound_next.play()
                    menu_option = (menu_option - 1) % 40
                    menu_level_info = ''
                elif event.key == K_d or event.key == K_RIGHT:
                    menu_option = (menu_option + 1) % 40
                    menu_level_info = ''
                    sound_next.play()
                elif event.key == K_RETURN:
                    if menu_option < unlocked_level:
                        map_level = int(menu_option + 1)
                        menu_level_select = False
                        menu_weapon_select = True
                        menu_option = players_bow.index if back_to_menu else weapon_index
                        current_menu = "weapon"
                    else:
                        menu_level_info = 'This level is not yet unlocked'
            # Select Weapon
            elif current_menu == "weapon":
                if event.key == K_a or event.key == K_LEFT:
                    menu_option = (menu_option - 1) % 4
                    menu_level_info = ''
                    if weapon_menu_sound:
                        sound_next.play()
                elif event.key == K_d or event.key == K_RIGHT:
                    menu_option = (menu_option + 1) % 4
                    menu_level_info = ''
                    if weapon_menu_sound:
                        sound_next.play()
                elif event.key == K_RETURN:
                    menu_weapon_select = False
                    weapon_menu_sound = False
                    gs.update_value('weapon', int(menu_option))
                    game_start = True
                    player.dead = True
                    main_menu = True
                    menu_about = False
                    players_bow.index = int(menu_option)
    # endregion

    # region 8.2 Intro and select menu
    if intro_running:
        display_surface.blit(intro_img, intro_rect)
        pygame.draw.rect(display_surface, gs.COLOR_01, (175, 285, load_bar_width, 55))

        if load_bar_width < 790:
            load_bar_width += 5
        else:
            intro_running = False
            menu_level_select = True

    if menu_level_select:
        pygame.mouse.set_visible(True)
        clear_display()

        # Draw and update cursor
        mouse_position = pygame.mouse.get_pos()

        menu_options = [str(number) for number in range(1, max_level + 1)]

        main_text = font_menu_big.render('Select Level', True, gs.WHITE)
        main_text_rect = main_text.get_rect(center=(gs.WINDOW_WIDTH // 2, 90))
        display_surface.blit(main_text, main_text_rect)

        info_text = font_menu_small.render(menu_level_info, True, gs.RED)
        info_text_rect = info_text.get_rect(center=(gs.WINDOW_WIDTH // 2, gs.WINDOW_HEIGHT - 90))
        display_surface.blit(info_text, info_text_rect)

        for i, option in enumerate(menu_options):
            text_color = gs.GREEN if i < unlocked_level else gs.RED
            frame_color = gs.COLOR_01 if i == menu_option else gs.WHITE
            frame_thickness = 6 if i == menu_option else 2
            text = font_title.render(option, True, text_color)
            text_rect = text.get_rect(center=(-55 + 120 * (2 + i % 8), 120 + 120 * (1 + i // 8)))
            display_surface.blit(text, text_rect)

            pygame.draw.rect(display_surface, frame_color, (120 * (2 + i % 8) - 100, 70 + 120 * (1 + i // 8), 90, 90),
                             frame_thickness)

    if menu_weapon_select:
        clear_display(show_info=False)
        menu_options = [str(number) for number in range(1, 5)]
        main_text = font_menu_big.render('Select Weapon', True, gs.WHITE)
        main_text_rect = main_text.get_rect(center=(gs.WINDOW_WIDTH // 2, 90))
        display_surface.blit(main_text, main_text_rect)

        for x in range(1, 5):
            frame_color = gs.COLOR_01 if x == menu_option else gs.WHITE
            for y in range(8):
                text_title = font_menu_small.render(weapon_value_list[x - 1][y], True, gs.WHITE)
                text_title_rect = text_title.get_rect(left=20 + 235 * x, centery=370 + 65 * y)
                display_surface.blit(text_title, text_title_rect)
        for i in range(8):
            text_title = font_menu_small.render(weapon_title_list[i], True, gs.COLORS[i])
            text_title_rect = text_title.get_rect(right=200, centery=370 + 65 * i)
            display_surface.blit(text_title, text_title_rect)

        COLOR_LIST = [gs.RED, gs.GREEN, gs.BLUE, gs.MIX]
        for i in range(4):
            frame_color = COLOR_LIST[i] if i == menu_option else gs.WHITE
            display_surface.blit(images_bow_list[i], (250 + 235 * i, 175))
            pygame.draw.rect(display_surface, frame_color, (225 + 235 * i, 160, 205, 700), 4)
    # endregion

    if game_start:
        # region 8.3 Create Game
        display_surface.fill('black')

        # Draw Map
        world.draw(display_surface)
        floor_group.draw(display_surface)

        for arrow in arrow_group:
            arrow.draw(display_surface)
        for enemy in enemy_list:
            enemy.draw(display_surface)

        # Blit images
        player.draw(display_surface)
        players_bow.draw(display_surface)

        # Update and draw the Game
        screen_scroll = player.move_camera()
        world.update(screen_scroll)
        player.move(players_bow, world.obstacle_list)

        next_map, shop_map = player.obstacle_collision(world.next_map, world.teleport_list, world.shop_list,
                                                       world.shop_map)
        player.update()

        # Check shop map
        if shop_map is True:
            start_intro = True
            shop_map = False
            world_data = load_next_map()
            # load map data from csv file
            with open(f'./config/maps/shop_map.csv', newline="") as csvfile:
                reader = csv.reader(csvfile, delimiter=',')
                for x, row in enumerate(reader):
                    for y, tile in enumerate(row):
                        world_data[x][y] = int(tile)
            # create map
            world = Map()
            world.data(tile_list, world_data, item_list, characters_list)

            # Create enemy
            enemy_list = world.mob_list
            for enemy in enemy_list:
                enemy_group.add(enemy)

            # create player with previous stats
            [last_health, last_score, last_coin, last_mana, last_armor, last_resist, last_velocity, last_max_hp] = [
                player.health, player.score, player.coin_amount, player.boost_lvl, player.armor, player.fire_resist,
                player.velocity_normal, player.max_hp]

            player = world.player
            [player.health, player.score, player.coin_amount, player.boost_lvl, player.armor, player.fire_resist,
             player.velocity_normal, player.max_hp] = [last_health, last_score, last_coin, last_mana, last_armor,
                                                       last_resist, last_velocity, last_max_hp]

            # Create SHOP items
            for item in world.item_list:
                item_group.add(item)

        # Check next map (next level)
        if next_map is True:
            pygame.mouse.set_visible(False)
            start_intro = True
            next_map = False
            map_level += 1
            if gs.get_value('max_level') < map_level:
                gs.update_value('max_level', map_level)
            world_data = load_next_map()
            # load map data from csv file
            with open(f'./config/maps/level_{map_level}.csv', newline="") as csvfile:
                reader = csv.reader(csvfile, delimiter=',')
                for x, row in enumerate(reader):
                    for y, tile in enumerate(row):
                        world_data[x][y] = int(tile)
            # create map
            world = Map()
            world.data(tile_list, world_data, item_list, characters_list)

            # create player with previous stats
            last_health, last_score, last_coin, last_mana, \
                last_armor, last_resist, last_velocity, last_max_hp = player.health, player.score, player.coin_amount, \
                player.boost_lvl, player.armor, player.fire_resist, player.velocity_normal, player.max_hp
            player = world.player
            player.health, player.score, player.coin_amount, player.boost_lvl, player.armor, player.fire_resist, \
                player.velocity_normal, player.max_hp = last_health, last_score, last_coin, last_mana, last_armor, \
                last_resist, last_velocity, last_max_hp
            players_bow.storage += 1 + math.ceil(map_level // 7)
            player.map_level = map_level

            # add extra coin to player
            if GameOverText:
                player.coin_amount += max(1, map_level // 5)

            # Create enemy
            enemy_list = world.mob_list
            for enemy in enemy_list:
                enemy_group.add(enemy)

            floor_list = world.floor_list
            for floor in floor_list:
                floor_group.add(floor)

            # Create objects
            for item in world.item_list:
                item_group.add(item)

        # Show intro
        if start_intro is True:
            if intro_fade.fade():
                start_intro = False
                intro_fade.counter = 0

        if menu_intro is True:
            if menu_fade.fade():
                menu_intro = False
                menu_fade.counter = 0

        # Create enemy and arrow
        for enemy in enemy_list:
            fireball = enemy.move(screen_scroll, player, world.obstacle_list)
            if fireball:
                fireball_group.add(fireball)

            enemy_drop = enemy.update()
            if enemy_drop:
                item_group.add(enemy_drop)

        arrow = players_bow.update(player, players_bow)
        if arrow:
            arrow_group.add(arrow)

        for arrow in arrow_group:
            results: tuple = arrow.update(enemy_list, world.obstacle_list, screen_scroll)
            damage, damage_coordinates, critical_strike, slow = results
            if damage:
                damage_text = mobs.DamageText(damage, damage_coordinates, critical_strike, slow)
                damage_text_group.add(damage_text)

        for fireball in fireball_group:
            fireball.draw(display_surface)
            fireball_obj = fireball.update(player, screen_scroll, world.obstacle_list)

        # Draw group
        item_group.draw(display_surface)
        arrow_group.draw(display_surface)
        # enemy_group.draw(display_surface)
        damage_text_group.draw(display_surface)

        # enemy_group.update()
        item_group.update(screen_scroll, player, players_bow)
        floor_group.update(screen_scroll, player, players_bow)
        damage_text_group.update(screen_scroll)

        # Draw HUD
        player.draw_hud(display_surface, players_bow)
        display_surface.blit(coin_images[coin_index], (235, 8))

        # Draw and update cursor
        mouse_position = pygame.mouse.get_pos()
        cursor_x = mouse_position[0] - 7.5
        cursor_y = mouse_position[1] - 7.5
        cursor_rect = (mouse_position[0], mouse_position[1], 20, 20)
        display_surface.blit(cursor_img, (cursor_x, cursor_y))

        # Draw player stats:
        if weapon_info:
            pygame.draw.rect(display_surface, gs.BLACK, (gs.WINDOW_WIDTH - 200, gs.WINDOW_HEIGHT - 260, 200, 260))
            pygame.draw.rect(display_surface, gs.WHITE, (gs.WINDOW_WIDTH - 200, gs.WINDOW_HEIGHT - 260, 200, 260), 2)
            temp_info = 'Acceleration' if players_bow.deceleration <= 0 else 'Deceleration'
            info_list = [f'DMG: {players_bow.min_damage} - {players_bow.max_damage}',
                         f'Attack Speed: {1000 / players_bow.shot_cooldown:.2f}',
                         f'{temp_info}: {math.fabs(players_bow.deceleration)}',
                         f'Critical: {players_bow.critical_chance} %',
                         f'Slow: {players_bow.slow_chance} %',
                         f'Reduction.: {players_bow.armor_penetration} %']
            for i in range(len(info_list)):
                text = font_info.render(info_list[i], True, gs.COLORS[i])
                text_rect = text.get_rect(left=gs.WINDOW_WIDTH - 190, centery=gs.WINDOW_HEIGHT - 235 + i * 42)
                display_surface.blit(text, text_rect)

        if player_info:
            frame_offset = 200 if weapon_info else 0
            pygame.draw.rect(display_surface, gs.BLACK, (gs.WINDOW_WIDTH - 170 - frame_offset, gs.WINDOW_HEIGHT - 170,
                                                         170, 180))
            pygame.draw.rect(display_surface, gs.COLOR_01,
                             (gs.WINDOW_WIDTH - 170 - frame_offset, gs.WINDOW_HEIGHT - 170,
                              170, 180), 2)
            info_list = [f'HP: {player.health} / {player.max_hp}', f'Armor: {player.armor}',
                         f'Fire Resist: {player.fire_resist}', f'Speed: {player.velocity:.2f}']
            for i in range(len(info_list)):
                text = font_info.render(info_list[i], True, gs.COLORS[i + 5])
                text_rect = text.get_rect(left=gs.WINDOW_WIDTH - 160 - frame_offset,
                                          centery=gs.WINDOW_HEIGHT - 145 + i * 42)
                display_surface.blit(text, text_rect)

        # Draw mobs HP
        for mob in enemy_list:
            if mob.rect.colliderect(cursor_rect):
                hf.draw_text(display_surface, f'{mob.health} / {mob.max_health}', 80, 8, font_mob_hp)

        # TEMP
        # for enemy in enemy_list:
        #     pygame.draw.rect(display_surface, 'red', enemy.rect, 1)
        #     pygame.draw.rect(display_surface, 'green', enemy.collision_rect, 1)
        # for item in item_group:
        #     pygame.draw.rect(display_surface, 'blue', item.rect, 1)

        # endregion

        if player.dead:
            # region 8.4 Game Over
            pygame.mouse.set_visible(True)
            discord_link_activ = True
            unlocked_level = gs.get_value('max_level')
            FPS = gs.FPS_MENU
            # Check new score record
            if player.score > 0 and save_score:
                ws.append([player.score, datetime.datetime.now().strftime("%Y-%m-%d"), players_bow.index])
                wb.save(excel_file_path)
                save_score = False

            if player.score > high_score:
                high_score = player.score
                gs.update_value('high_score', player.score)
            start_intro = True
            # endregion

            # region 8.5 About Menu
            if menu_about:
                pygame.mouse.set_visible(True)
                clear_display()
                back_click = True
                main_menu = False
                menu_score = False
                menu_how_to_play = False
                menu_mob_info = False
                menu_item_info = False
                hf.draw_image(display_surface, menu_about_img, 0, 0)

                if button_info_exit.draw(display_surface):
                    game_running = False
                elif button_about_back.draw(display_surface):
                    main_menu = True
                    menu_about = False
                elif button_info_game.draw(display_surface):
                    menu_how_to_play = True
                    menu_about = False
                elif button_info_mob.draw(display_surface):
                    menu_mob_info = True
                    menu_about = False
                elif button_info_item.draw(display_surface):
                    menu_item_info = True
                    menu_about = False
                elif button_info_weapon.draw(display_surface):
                    menu_weapon_info = True
                    menu_about = False

            #  Draw About SubMenu 1
            if menu_how_to_play:
                # pygame.mouse.set_visible(True)
                clear_display()
                # hf.draw_image(display_surface, hf.scale_image(keyboard_img, 0.7), 6, 550)

                for i in range(len(HTP_info_text_list[menu_HTP_index])):
                    text_info = font_menu_item.render(HTP_info_text_list[menu_HTP_index][i], True, gs.WHITE)
                    text_info_rect = text_info.get_rect(center=(gs.WINDOW_WIDTH // 2, 60 + i * 50))
                    display_surface.blit(text_info, text_info_rect)

                hf.draw_image(display_surface, menu_how_to_play_images_list[menu_HTP_index], 312, 210)
                pygame.draw.rect(display_surface, 'yellow', (312, 210, 575, 485), 2)

                if button_next_option.draw(display_surface):
                    menu_HTP_index = (menu_HTP_index + 1) % 6
                if button_prev_option.draw(display_surface):
                    menu_HTP_index = (menu_HTP_index - 1) % 6
                if button_info_exit.draw(display_surface):
                    game_running = False
                elif button_info_back.draw(display_surface):
                    menu_how_to_play = False
                    menu_about = True

            #  Draw About SubMenu 2
            if menu_mob_info:
                clear_display()
                # pygame.mouse.set_visible(True)

                # Display text with mob stats
                for i in range(len(mob_info_text_list)):
                    text_info = font_menu_item.render(mob_info_text_list[i], True, gs.WHITE)
                    text_info_rect = text_info.get_rect(right=gs.WINDOW_WIDTH // 2 - 25, centery=350 + i * 50)
                    display_surface.blit(text_info, text_info_rect)
                    text_stats = font_menu_item.render(mob_stats_text_list[menu_mob_index][i], True, gs.WHITE)
                    text_stats_rect = text_stats.get_rect(left=gs.WINDOW_WIDTH // 2 + 25, centery=350 + i * 50)
                    display_surface.blit(text_stats, text_stats_rect)

                # Display text with extra mob info
                for i in range(len(mob_extra_info_list[menu_mob_index])):
                    text_extra_info = font_menu_item.render(mob_extra_info_list[menu_mob_index][i], True, gs.WHITE)
                    text_extra_info_rect = text_extra_info.get_rect(right=gs.WINDOW_WIDTH - 100, centery=100 + i * 50)
                    display_surface.blit(text_extra_info, text_extra_info_rect)

                # Display mob images
                hf.draw_image(display_surface, mob_menu_list[menu_mob_index + 1][1][mob_img_index],
                              100 + mob_img_pos_x, 100 + mob_img_pos_x)

                # Update mob images
                timer += 1
                if timer % 2 == 0:
                    mob_img_index = (mob_img_index + 1) % 3
                if timer % FPS == 0:
                    mob_img_pos_x += 1

                # Draw button
                if button_next_option.draw(display_surface):
                    menu_mob_index = (menu_mob_index + 1) % 9
                    mob_img_pos_x = 0
                if button_prev_option.draw(display_surface):
                    menu_mob_index = (menu_mob_index - 1) % 9
                    mob_img_pos_x = 0
                if button_info_exit.draw(display_surface):
                    game_running = False
                elif button_info_back.draw(display_surface):
                    menu_mob_info = False
                    menu_about = True
                    mob_img_pos_x = 0

            #  Draw About SubMenu 3
            if menu_item_info:
                clear_display()
                display_surface.blit(potion_images_list[menu_item_index], (500, 200))
                for i in range(len(potion_info_text_list[menu_item_index])):
                    text = font_menu_small.render(potion_info_text_list[menu_item_index][i], True, gs.WHITE)
                    text_rect = text.get_rect(center=(gs.WINDOW_WIDTH // 2, gs.WINDOW_HEIGHT // 2 + 100 + 60 * i))
                    display_surface.blit(text, text_rect)

                if button_next_option.draw(display_surface):
                    menu_item_index = (menu_item_index + 1) % 13
                if button_prev_option.draw(display_surface):
                    menu_item_index = (menu_item_index - 1) % 13
                if button_info_exit.draw(display_surface):
                    game_running = False
                elif button_info_back.draw(display_surface):
                    menu_item_info = False
                    menu_about = True

            if menu_weapon_info:
                clear_display()
                back_click = True
                main_menu = False

                # Draw button
                if button_info_exit.draw(display_surface):
                    game_running = False
                elif button_score_back.draw(display_surface):
                    menu_weapon_info = False
                    main_menu = True
                elif button_next_option.draw(display_surface):
                    menu_weapon_info_index = (menu_weapon_info_index + 1) % len(weapon_title_list_full)
                elif button_prev_option.draw(display_surface):
                    menu_weapon_info_index = (menu_weapon_info_index - 1) % len(weapon_title_list_full)
                elif button_next_weapon.draw(display_surface):
                    weapon_index = (weapon_index + 1) % 4
                elif button_prev_weapon.draw(display_surface):
                    weapon_index = (weapon_index - 1) % 4

                # Draw basic info
                for i in range(len(weapon_title_list_full)):
                    hf.draw_text(surface=display_surface, text=weapon_title_list_full[i], x=350, y=40 + 55 * i,
                                 font=font_menu_item, color=gs.COLORS[i])

                for i in range(len(weapon_value_list[1])):
                    hf.draw_text(surface=display_surface, text=weapon_value_list[weapon_index][i], x=690,
                                 y=40 + 55 * i, font=font_menu_item, color=gs.COLORS[i])

                # Draw extra info
                # Display text with extra mob info
                pygame.draw.rect(display_surface, gs.COLORS[menu_weapon_info_index],
                                 (330, 40 + 55 * menu_weapon_info_index, 520, 50), 2)

                for i in range(len(weapon_info_list_full[menu_weapon_info_index])):
                    text_extra_info = font_menu_item.render(weapon_info_list_full[menu_weapon_info_index][i],
                                                            True, gs.WHITE)
                    text_extra_info_rect = text_extra_info.get_rect(center=(gs.WINDOW_WIDTH // 2, 625 + 55 * i))
                    display_surface.blit(text_extra_info, text_extra_info_rect)

                # Draw BOW image
                hf.draw_image(surface=display_surface, image=images_bow_list[weapon_index], x=950, y=150)
            # endregion

            # 8.6 Score Menu
            if menu_score:
                clear_display()
                back_click = True
                main_menu = False
                hf.draw_image(display_surface, high_score_img, 800, 250)
                hf.create_text(display_surface, f"Last try: {player.score:03d}", x=900, y=510,
                               size=45, font='Chiller', color=gs.WHITE)
                hf.create_text(display_surface, f"{high_score:04d}", x=930, y=330,
                               size=85, font='Chiller', color=gs.MIX)
                # Draw button
                if button_info_exit.draw(display_surface):
                    game_running = False
                elif button_score_back.draw(display_surface):
                    menu_score = False
                    main_menu = True

                if not restart_score:
                    if button_score_restart.draw(display_surface):
                        restart_score = True
                        show_top11 = False
                        hf.rename_file(old_name='score', new_name='score_backUp')

                # Load excel data - top 11
                if show_top11:
                    wb = openpyxl.load_workbook(excel_file_path)
                    ws = wb.active
                    sorted_results = sorted(ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row),
                                            key=lambda var_x: var_x[0].value, reverse=True)
                    top10_results = sorted_results[:11]
                    # Draw Best score
                    for i, result in enumerate(top10_results):
                        text_top3 = f"Score: {result[0].value}, Date: {result[1].value}"
                        hf.create_text(display_surface, f"{i + 1}.", x=200,
                                       y=35 + 65 * i - math.ceil(3.5 * i // 2 + 12.5 * i // 3),
                                       size=math.ceil(50 - 1.75 * i), font='Chiller',
                                       random_color=False, color=gs.COLORS[i])
                        hf.create_text(display_surface, text_top3, x=340,
                                       y=35 + 65 * i - math.ceil(3.5 * i // 2 + 12.5 * i // 3),
                                       size=math.ceil(50 - 1.75 * i), font='Chiller',
                                       random_color=False, color=gs.COLORS[i])
                        hf.draw_image(surface=display_surface,
                                      image=hf.scale_size_image(images_bow_list[int(result[2].value)],
                                                                math.ceil(70 - 2.65 * i), math.ceil(60 - 2.65 * i)),
                                      x=250, y=35 + 65 * i - math.ceil(3.5 * i // 2 + 12.5 * i // 3))

            # region 8.7 Main Menu
            if main_menu:
                clear_display()
                hf.draw_image(display_surface, archer_img, 850, 350)
                # OFF previous menu
                menu_about = False
                menu_how_to_play = False
                menu_mob_info = False
                menu_item_info = False
                menu_weapon_info = False
                menu_score = False
                restart_score = False
                show_top11 = True

                # Show GameOverScreen
                if GameOverText:
                    GAMEOVER_text = font_big.render(' GAME OVER', True, gs.RED)
                    game_over_text = font_title.render(f"FINAL SCORE: {player.score}", True, gs.WHITE)
                else:
                    game_over_text = font_title.render(f"HIGH SCORE: {high_score}", True, gs.COLORS[3])
                    GAMEOVER_text = font_big.render('Dungeon Archer', True, gs.COLORS[2])

                display_surface.blit(GAMEOVER_text, GAMEOVER_rect)
                display_surface.blit(game_over_text, game_over_rect)

                # Draw Main Menu Buttons
                if button_restart.draw(display_surface):
                    pygame.mouse.set_visible(False)
                    FPS = gs.FPS_GAME
                    discord_link_activ = False
                    start_intro = True
                    next_map = False
                    GameOverText = True
                    save_score = True
                    world_data = load_next_map()
                    with open(f'./config/maps/level_{map_level}.csv', newline="") as csvfile:
                        reader = csv.reader(csvfile, delimiter=',')
                        for x, row in enumerate(reader):
                            for y, tile in enumerate(row):
                                if 0 <= x < len(world_data) and 0 <= y < len(world_data[x]):
                                    world_data[x][y] = int(tile)
                                else:
                                    print(f'Error! x:{x}, y:{y} out of range, change game_setting.py')
                    world = Map()
                    world.data(tile_list, world_data, item_list, characters_list)
                    player = world.player

                    # player restart values
                    player.restart()
                    player.map_level = map_level

                    # restart weapon values
                    players_bow.restart()

                    enemy_list = world.mob_list
                    for enemy in enemy_list:
                        enemy_group.add(enemy)
                    for item in world.item_list:
                        item_group.add(item)
                    for floor in world.floor_list:
                        floor_group.add(floor)

                elif button_about.draw(display_surface):
                    menu_about = True

                elif button_main_menu.draw(display_surface):
                    if back_click:
                        back_click = False
                    else:
                        back_to_menu = True
                        menu_option = map_level - 1
                        game_start = False
                        menu_level_select = True
                        current_menu = "level"

                elif button_exit.draw(display_surface):
                    game_running = False

                elif button_score.draw(display_surface):
                    menu_score = True
            # endregion

    # Update display - update coin images in HUD
    timer += 1
    if timer % 9 == 0:
        coin_index = (coin_index + 1) % 5

    # update music player
    if button_music.draw(display_surface):
        change_music_level()
        button_music.update()

    # Clock settings
    pygame.display.update()
    clock.tick(FPS)

pygame.quit()
